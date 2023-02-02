from typing import Optional, Type
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.db import db
from sqlalchemy.sql import text, desc
from app.models import Admin, User, Order


def get_entity_by_id(
    Entity: Type[User | Admin], entity_id: Optional[str | int]
) -> Optional[User | Admin]:
    return db.session.get(Entity, entity_id)


def create_admin(username: str, password: str) -> None:
    admin = Admin(username=username, password=password)
    db.session.add(admin)


def get_admin_by_username(username: str) -> Optional[Admin]:
    stmt = db.select(Admin).where(Admin.username == username)
    result = db.session.execute(stmt)
    admin = result.scalar()
    return admin


def create_user(
    gender: str, age: str, zip_code: str,
    marital_status: str, income: str
) -> User:
    user = User(
        gender=gender,
        age=age,
        zip_code=zip_code,
        marital_status=marital_status,
        income=income
    )
    db.session.add(user)
    return user


def update_user(
    user_id: str, gender: str, age: str,
    zip_code: str, marital_status: str, income: str
) -> Optional[Type[User]]:
    user = get_entity_by_id(User, user_id)

    if user is not None:
        user.gender = gender
        user.age = age
        user.zip_code = zip_code
        user.marital_status = marital_status
        user.income = income
    return user


def get_users() -> list[User]:
    stmt = db.select(User).order_by(desc(User.id)).options(
        db.selectinload(User.orders)
    )
    result = db.session.execute(stmt)
    users = result.scalars().all()
    return users


def create_order(
    user_id: str, email: str, price: str, date: str, page: str
) -> None:
    order = Order(
        email=email,
        price=price,
        date=date,
        page=page
    )
    db.session.add(order)

    user = get_entity_by_id(User, user_id)
    if user is not None:
        user.orders.append(order)


def delete_tables() -> None:
    delete_stmts = (
        text("DELETE FROM users"),
        text("DELETE FROM orders"),
        text("TRUNCATE users, orders, user_order RESTART IDENTITY"),
    )
    for stmt in delete_stmts:
        db.session.execute(stmt)


def export_to_excel() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика"
    users = get_users()

    row = 1
    for user in users:
        for col, user_column in enumerate(user.__table__.columns.keys()):
            ws.cell(column=col+1, row=row, value=getattr(user, user_column))
        row += 1

        for order in user.orders:
            for col, order_column in enumerate(order.__table__.columns.keys()):
                if order_column != 'id':
                    ws.cell(
                        column=col+1, row=row,
                        value=getattr(order, order_column)
                    )
            row += 1
        row += 1

    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].auto_size = True
    return wb


def export_statistics(filename: str) -> None:
    wb = export_to_excel()
    wb.save(f"uploads/{filename}")
